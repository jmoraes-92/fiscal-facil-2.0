from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Header
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr
from typing import Optional, List
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
from utils.auth import verify_password, get_password_hash, create_access_token, decode_token
from utils.brasil_api import consultar_cnpj
from utils.xml_parser import parse_xml_nota

load_dotenv()

# Configura칞칚o MongoDB
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
client = AsyncIOMotorClient(MONGO_URL)
db = client.fiscal_facil

# FastAPI App
app = FastAPI(title="Fiscal F치cil API", version="2.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== SCHEMAS ====================
class UsuarioRegistro(BaseModel):
    nome: str
    email: EmailStr
    senha: str
    telefone: Optional[str] = None

class UsuarioLogin(BaseModel):
    email: EmailStr
    senha: str

class CnaePermitido(BaseModel):
    cnae_codigo: str
    codigo_servico_municipal: str
    descricao: Optional[str] = None

class EmpresaCadastro(BaseModel):
    cnpj: str
    razao_social: str
    nome_fantasia: Optional[str] = None
    regime_tributario: str
    data_abertura: Optional[str] = None
    cnaes_permitidos: List[CnaePermitido]

# ==================== AUTH MIDDLEWARE ====================
async def get_current_user(authorization: Optional[str] = Header(None)):
    from bson import ObjectId
    
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token n칚o fornecido")
    
    token = authorization.replace("Bearer ", "")
    payload = decode_token(token)
    
    if not payload:
        raise HTTPException(status_code=401, detail="Token inv치lido ou expirado")
    
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inv치lido")
    
    try:
        user = await db.usuarios.find_one({"_id": ObjectId(user_id)})
    except:
        raise HTTPException(status_code=401, detail="Token inv치lido")
    
    if not user:
        raise HTTPException(status_code=401, detail="Usu치rio n칚o encontrado")
    
    return user

# ==================== ROTAS DE AUTENTICA칂츾O ====================
@app.post("/api/auth/registro")
async def registrar_usuario(usuario: UsuarioRegistro):
    # Verifica se o email j치 existe
    existe = await db.usuarios.find_one({"email": usuario.email})
    if existe:
        raise HTTPException(status_code=400, detail="Email j치 cadastrado")
    
    # Cria o usu치rio
    usuario_doc = {
        "nome": usuario.nome,
        "email": usuario.email,
        "senha_hash": get_password_hash(usuario.senha),
        "telefone": usuario.telefone,
        "data_criacao": datetime.utcnow().isoformat()
    }
    
    result = await db.usuarios.insert_one(usuario_doc)
    usuario_id = str(result.inserted_id)
    
    # Gera token
    access_token = create_access_token(data={"sub": usuario_id})
    
    return {
        "mensagem": "Usu치rio cadastrado com sucesso",
        "access_token": access_token,
        "token_type": "bearer",
        "usuario": {
            "id": usuario_id,
            "nome": usuario.nome,
            "email": usuario.email
        }
    }

@app.post("/api/auth/login")
async def login(credenciais: UsuarioLogin):
    # Busca usu치rio
    usuario = await db.usuarios.find_one({"email": credenciais.email})
    if not usuario:
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    # Verifica senha
    if not verify_password(credenciais.senha, usuario["senha_hash"]):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    # Gera token
    usuario_id = str(usuario["_id"])
    access_token = create_access_token(data={"sub": usuario_id})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "usuario": {
            "id": usuario_id,
            "nome": usuario["nome"],
            "email": usuario["email"]
        }
    }

@app.get("/api/auth/me")
async def obter_usuario_atual(current_user: dict = Depends(get_current_user)):
    return {
        "id": str(current_user["_id"]),
        "nome": current_user["nome"],
        "email": current_user["email"],
        "telefone": current_user.get("telefone")
    }

# ==================== ROTAS DE EMPRESAS ====================
@app.get("/api/empresas/consulta/{cnpj}")
async def consultar_cnpj_endpoint(cnpj: str, current_user: dict = Depends(get_current_user)):
    return consultar_cnpj(cnpj)

@app.post("/api/empresas")
async def cadastrar_empresa(empresa: EmpresaCadastro, current_user: dict = Depends(get_current_user)):
    usuario_id = str(current_user["_id"])
    cnpj_limpo = "".join([n for n in empresa.cnpj if n.isdigit()])
    
    # Verifica se j치 existe
    existe = await db.empresas.find_one({"cnpj": cnpj_limpo})
    if existe:
        raise HTTPException(status_code=400, detail="Empresa j치 cadastrada")
    
    # Cria a empresa
    empresa_doc = {
        "usuario_id": usuario_id,
        "cnpj": cnpj_limpo,
        "razao_social": empresa.razao_social,
        "nome_fantasia": empresa.nome_fantasia,
        "regime_tributario": empresa.regime_tributario,
        "data_abertura": empresa.data_abertura,
        "cnaes_permitidos": [cnae.dict() for cnae in empresa.cnaes_permitidos],
        "data_cadastro": datetime.utcnow().isoformat()
    }
    
    result = await db.empresas.insert_one(empresa_doc)
    
    return {
        "mensagem": "Empresa cadastrada com sucesso",
        "id": str(result.inserted_id)
    }

@app.get("/api/empresas")
async def listar_empresas(current_user: dict = Depends(get_current_user)):
    usuario_id = str(current_user["_id"])
    empresas = []
    
    async for empresa in db.empresas.find({"usuario_id": usuario_id}):
        empresas.append({
            "id": str(empresa["_id"]),
            "cnpj": empresa["cnpj"],
            "razao_social": empresa["razao_social"],
            "nome_fantasia": empresa.get("nome_fantasia"),
            "regime_tributario": empresa["regime_tributario"]
        })
    
    return empresas

@app.get("/api/empresas/{empresa_id}")
async def obter_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    empresa["id"] = str(empresa.pop("_id"))
    return empresa

# ==================== ROTAS DE NOTAS FISCAIS ====================

# Fun칞칚o auxiliar para processar um 칰nico XML
async def processar_xml_nota(empresa_id: str, empresa: dict, conteudo: bytes, nome_arquivo: str):
    """
    Processa um arquivo XML e retorna o resultado do processamento.
    N칚o levanta exce칞칫es, retorna dict com sucesso ou erro.
    """
    try:
        # Parse do XML
        dados_xml = parse_xml_nota(conteudo)
        
        if "erro" in dados_xml:
            return {
                "sucesso": False,
                "nome_arquivo": nome_arquivo,
                "erro": dados_xml["erro"]
            }
        
        # Auditoria: Verifica se o c칩digo de servi칞o est치 permitido
        codigo_servico = dados_xml['codigo_servico']
        cnaes_permitidos = empresa.get("cnaes_permitidos", [])
        
        cnae_encontrado = None
        for cnae in cnaes_permitidos:
            if cnae.get("codigo_servico_municipal") == codigo_servico:
                cnae_encontrado = cnae
                break
        
        status = "APROVADA"
        mensagem = "Nota fiscal em conformidade"
        
        if not cnae_encontrado:
            status = "ERRO_CNAE"
            mensagem = f"C칩digo de servi칞o '{codigo_servico}' n칚o autorizado para este CNPJ"
        
        # Salva a nota (incluindo o XML original)
        nota_doc = {
            "empresa_id": empresa_id,
            "numero_nota": dados_xml['numero_nota'],
            "data_emissao": dados_xml['data_emissao'],
            "chave_validacao": dados_xml.get('chave_validacao'),
            "cnpj_tomador": dados_xml.get('cnpj_tomador'),
            "codigo_servico_utilizado": codigo_servico,
            "valor_total": dados_xml['valor_total'],
            "status_auditoria": status,
            "mensagem_erro": mensagem,
            "xml_original": dados_xml.get('xml_bruto', ''),  # Armazena o XML completo
            "data_importacao": datetime.utcnow().isoformat()
        }
        
        result = await db.notas_fiscais.insert_one(nota_doc)
        
        return {
            "sucesso": True,
            "nome_arquivo": nome_arquivo,
            "nota": {
                "id": str(result.inserted_id),
                "numero_nota": nota_doc["numero_nota"],
                "status_auditoria": nota_doc["status_auditoria"],
                "valor_total": nota_doc["valor_total"]
            }
        }
        
    except Exception as e:
        return {
            "sucesso": False,
            "nome_arquivo": nome_arquivo,
            "erro": f"Erro ao processar: {str(e)}"
        }

@app.post("/api/notas/importar/{empresa_id}")
async def importar_nota_xml(
    empresa_id: str, 
    file: UploadFile = File(...), 
    current_user: dict = Depends(get_current_user)
):
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # L칡 o XML
    conteudo = await file.read()
    resultado = await processar_xml_nota(empresa_id, empresa, conteudo, file.filename)
    
    if not resultado["sucesso"]:
        raise HTTPException(status_code=400, detail=resultado["erro"])
    
    # Busca a nota completa para retornar
    from bson import ObjectId
    nota = await db.notas_fiscais.find_one({"_id": ObjectId(resultado["nota"]["id"])})
    
    return {
        "id": str(nota["_id"]),
        "numero_nota": nota["numero_nota"],
        "data_emissao": nota["data_emissao"],
        "codigo_servico_utilizado": nota["codigo_servico_utilizado"],
        "valor_total": nota["valor_total"],
        "status_auditoria": nota["status_auditoria"],
        "mensagem_erro": nota["mensagem_erro"],
        "chave_validacao": nota.get("chave_validacao"),
        "cnpj_tomador": nota.get("cnpj_tomador"),
        "data_importacao": nota["data_importacao"]
    }

@app.post("/api/notas/importar-lote/{empresa_id}")
async def importar_notas_em_lote(
    empresa_id: str,
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Importa m칰ltiplos arquivos XML de notas fiscais de uma vez.
    Processa todos os arquivos de forma "graceful" - se um falhar, continua os outros.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Valida quantidade de arquivos (m치ximo 100 por upload)
    if len(files) > 100:
        raise HTTPException(status_code=400, detail="M치ximo de 100 arquivos por upload")
    
    # Processa cada arquivo
    resultados = []
    sucessos = 0
    falhas = 0
    detalhes_falhas = []
    
    for idx, file in enumerate(files, 1):
        # L칡 o conte칰do
        conteudo = await file.read()
        
        # Processa o XML
        resultado = await processar_xml_nota(empresa_id, empresa, conteudo, file.filename)
        
        if resultado["sucesso"]:
            sucessos += 1
        else:
            falhas += 1
            detalhes_falhas.append({
                "arquivo": file.filename,
                "erro": resultado["erro"]
            })
        
        resultados.append(resultado)
    
    # Retorna resumo
    return {
        "total_arquivos": len(files),
        "sucesso": sucessos,
        "falhas": falhas,
        "detalhes_falhas": detalhes_falhas,
        "resultados": resultados
    }

@app.get("/api/notas/{nota_id}/detalhes")
async def obter_detalhes_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Retorna todos os detalhes de uma nota fiscal, incluindo XML original.
    """
    from bson import ObjectId
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inv치lido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota n칚o encontrada")
    
    # Verifica se a empresa da nota pertence ao usu치rio
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Retorna nota com todos os dados
    return {
        "id": str(nota["_id"]),
        "numero_nota": nota.get("numero_nota"),
        "data_emissao": nota.get("data_emissao"),
        "chave_validacao": nota.get("chave_validacao"),
        "cnpj_tomador": nota.get("cnpj_tomador"),
        "codigo_servico_utilizado": nota.get("codigo_servico_utilizado"),
        "valor_total": nota.get("valor_total"),
        "status_auditoria": nota.get("status_auditoria"),
        "mensagem_erro": nota.get("mensagem_erro"),
        "xml_original": nota.get("xml_original", ''),
        "data_importacao": nota.get("data_importacao"),
        "empresa": {
            "razao_social": empresa.get("razao_social"),
            "cnpj": empresa.get("cnpj"),
            "regime_tributario": empresa.get("regime_tributario")
        }
    }

@app.get("/api/notas/empresa/{empresa_id}")
async def listar_notas_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    # Verifica se a empresa pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Lista as notas com c치lculo de imposto estimado
    notas = []
    async for nota in db.notas_fiscais.find({"empresa_id": empresa_id}):
        valor_total = nota.get("valor_total", 0)
        
        # C치lculo de imposto estimado (Anexo III - 6%)
        imposto_estimado = valor_total * 0.06
        
        notas.append({
            "id": str(nota["_id"]),
            "numero_nota": nota["numero_nota"],
            "data_emissao": nota["data_emissao"],
            "codigo_servico_utilizado": nota["codigo_servico_utilizado"],
            "valor_total": valor_total,
            "imposto_estimado": round(imposto_estimado, 2),  # NOVO
            "status_auditoria": nota["status_auditoria"],
            "mensagem_erro": nota.get("mensagem_erro"),
            "data_importacao": nota["data_importacao"]
        })
    
    return notas

@app.get("/api/notas/estatisticas/{empresa_id}")
async def obter_estatisticas(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    # Verifica acesso
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID inv치lido")
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Estat칤sticas
    total = await db.notas_fiscais.count_documents({"empresa_id": empresa_id})
    aprovadas = await db.notas_fiscais.count_documents({"empresa_id": empresa_id, "status_auditoria": "APROVADA"})
    erros = await db.notas_fiscais.count_documents({"empresa_id": empresa_id, "status_auditoria": {"$ne": "APROVADA"}})
    
    # Valor total
    pipeline = [
        {"$match": {"empresa_id": empresa_id}},
        {"$group": {"_id": None, "total": {"$sum": "$valor_total"}}}
    ]
    
    resultado = await db.notas_fiscais.aggregate(pipeline).to_list(1)
    valor_total = resultado[0]["total"] if resultado else 0
    
    return {
        "total_notas": total,
        "aprovadas": aprovadas,
        "com_erros": erros,
        "valor_total": valor_total
    }

@app.delete("/api/notas/{nota_id}")
async def excluir_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Exclui uma nota fiscal. Verifica se a nota pertence a uma empresa do usu치rio.
    """
    from bson import ObjectId
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inv치lido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota n칚o encontrada")
    
    # Verifica se a empresa da nota pertence ao usu치rio
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Exclui a nota
    await db.notas_fiscais.delete_one({"_id": ObjectId(nota_id)})
    
    return {
        "mensagem": "Nota exclu칤da com sucesso",
        "nota_id": nota_id
    }

@app.put("/api/empresas/{empresa_id}")
async def atualizar_empresa(
    empresa_id: str,
    dados: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Atualiza dados de uma empresa. Permite editar raz칚o social, nome fantasia,
    regime tribut치rio e CNAEs permitidos.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Campos que podem ser atualizados
    campos_permitidos = [
        "razao_social",
        "nome_fantasia",
        "regime_tributario",
        "cnaes_permitidos",
        "limite_faturamento_anual"
    ]
    
    # Prepara o update
    update_data = {}
    for campo in campos_permitidos:
        if campo in dados:
            update_data[campo] = dados[campo]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo v치lido para atualizar")
    
    # Atualiza no banco
    await db.empresas.update_one(
        {"_id": ObjectId(empresa_id)},
        {"$set": update_data}
    )
    
    # Retorna empresa atualizada
    empresa_atualizada = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    empresa_atualizada["id"] = str(empresa_atualizada.pop("_id"))
    
    return {
        "mensagem": "Empresa atualizada com sucesso",
        "empresa": empresa_atualizada
    }

@app.delete("/api/empresas/{empresa_id}")
async def excluir_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """
    Exclui uma empresa e todas as suas notas fiscais associadas.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Conta quantas notas ser칚o exclu칤das
    total_notas = await db.notas_fiscais.count_documents({"empresa_id": empresa_id})
    
    # Exclui todas as notas da empresa
    await db.notas_fiscais.delete_many({"empresa_id": empresa_id})
    
    # Exclui a empresa
    await db.empresas.delete_one({"_id": ObjectId(empresa_id)})
    
    return {
        "mensagem": "Empresa e suas notas exclu칤das com sucesso",
        "empresa_id": empresa_id,
        "notas_excluidas": total_notas
    }

# ==================== RELAT칍RIOS ====================
@app.get("/api/notas/{nota_id}/pdf")
async def gerar_pdf_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Gera um PDF formatado da nota fiscal para visualiza칞칚o.
    """
    from bson import ObjectId
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inv치lido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota n칚o encontrada")
    
    # Verifica se a empresa da nota pertence ao usu치rio
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Cria o PDF em mem칩ria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=10
    )
    
    # T칤tulo
    elements.append(Paragraph("NOTA FISCAL DE SERVI칂OS ELETR칎NICA", title_style))
    elements.append(Spacer(1, 10*mm))
    
    # Informa칞칫es da Empresa Prestadora
    elements.append(Paragraph("DADOS DA EMPRESA PRESTADORA", header_style))
    
    empresa_data = [
        ['Raz칚o Social:', empresa.get('razao_social', 'N/A')],
        ['CNPJ:', empresa.get('cnpj', 'N/A')],
        ['Regime Tribut치rio:', empresa.get('regime_tributario', 'N/A')]
    ]
    
    empresa_table = Table(empresa_data, colWidths=[40*mm, 130*mm])
    empresa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(empresa_table)
    elements.append(Spacer(1, 8*mm))
    
    # Dados da Nota Fiscal
    elements.append(Paragraph("DADOS DA NOTA FISCAL", header_style))
    
    # Formata data
    data_emissao = nota.get('data_emissao', '')
    try:
        data_obj = datetime.fromisoformat(data_emissao.replace('Z', '+00:00'))
        data_formatada = data_obj.strftime('%d/%m/%Y 맙 %H:%M')
    except:
        data_formatada = data_emissao
    
    # Status com cor
    status = nota.get('status_auditoria', 'N/A')
    status_color = colors.green if status == 'APROVADA' else colors.red
    
    nota_data = [
        ['N칰mero da Nota:', str(nota.get('numero_nota', 'N/A'))],
        ['Data de Emiss칚o:', data_formatada],
        ['Chave de Valida칞칚o:', nota.get('chave_validacao', 'N/A')],
        ['C칩digo de Servi칞o:', nota.get('codigo_servico_utilizado', 'N/A')],
        ['CNPJ Tomador:', nota.get('cnpj_tomador', 'N/A')],
        ['Valor Total:', f"R$ {nota.get('valor_total', 0):.2f}"]
    ]
    
    nota_table = Table(nota_data, colWidths=[40*mm, 130*mm])
    nota_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(nota_table)
    elements.append(Spacer(1, 8*mm))
    
    # Status da Auditoria
    elements.append(Paragraph("RESULTADO DA AUDITORIA", header_style))
    
    auditoria_data = [
        ['Status:', status],
        ['Resultado:', nota.get('mensagem_erro', 'Nota fiscal em conformidade')]
    ]
    
    auditoria_table = Table(auditoria_data, colWidths=[40*mm, 130*mm])
    auditoria_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#dcfce7') if status == 'APROVADA' else colors.HexColor('#fee2e2')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('TEXTCOLOR', (1, 0), (1, 0), status_color),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(auditoria_table)
    elements.append(Spacer(1, 8*mm))
    
    # Rodap칠
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    data_importacao = nota.get('data_importacao', '')
    try:
        data_import_obj = datetime.fromisoformat(data_importacao.replace('Z', '+00:00'))
        data_import_formatada = data_import_obj.strftime('%d/%m/%Y 맙 %H:%M')
    except:
        data_import_formatada = data_importacao
    
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"Documento gerado em {datetime.utcnow().strftime('%d/%m/%Y 맙 %H:%M')}", footer_style))
    elements.append(Paragraph(f"Importado em: {data_import_formatada}", footer_style))
    elements.append(Paragraph("Fiscal F치cil - Sistema de Auditoria Fiscal", footer_style))
    
    # Gera o PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Retorna como streaming response
    filename = f"nota_{nota.get('numero_nota', 'fiscal')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

@app.get("/api/relatorios/inconsistencias/{empresa_id}")
async def gerar_relatorio_inconsistencias(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """
    Gera um relat칩rio Excel com todas as notas que possuem inconsist칡ncias/erros.
    """
    from bson import ObjectId
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Verifica se a empresa pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Busca apenas notas com erros
    notas_com_erro = []
    async for nota in db.notas_fiscais.find({
        "empresa_id": empresa_id,
        "status_auditoria": {"$ne": "APROVADA"}
    }).sort("data_emissao", -1):
        notas_com_erro.append(nota)
    
    if not notas_com_erro:
        raise HTTPException(
            status_code=404, 
            detail="Nenhuma inconsist칡ncia encontrada. Todas as notas est칚o aprovadas!"
        )
    
    # Cria o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inconsist칡ncias"
    
    # Cabe칞alho do relat칩rio
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Relat칩rio de Inconsist칡ncias - {empresa.get('razao_social', '')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"CNPJ: {empresa.get('cnpj', '')}"
    ws['A3'] = f"Data do Relat칩rio: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}"
    ws['A4'] = f"Total de Inconsist칡ncias: {len(notas_com_erro)}"
    
    # Cabe칞alhos da tabela
    headers = ['N칰mero da Nota', 'Data de Emiss칚o', 'C칩digo de Servi칞o', 'Valor (R$)', 'Status', 'Erro Encontrado']
    header_row = 6
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row_idx, nota in enumerate(notas_com_erro, header_row + 1):
        ws.cell(row=row_idx, column=1, value=nota.get('numero_nota'))
        
        # Data de emiss칚o formatada
        data_str = nota.get('data_emissao', '')
        if data_str:
            try:
                data_obj = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
                ws.cell(row=row_idx, column=2, value=data_obj.strftime('%d/%m/%Y'))
            except:
                ws.cell(row=row_idx, column=2, value=data_str)
        
        ws.cell(row=row_idx, column=3, value=nota.get('codigo_servico_utilizado'))
        ws.cell(row=row_idx, column=4, value=nota.get('valor_total'))
        ws.cell(row=row_idx, column=5, value=nota.get('status_auditoria'))
        ws.cell(row=row_idx, column=6, value=nota.get('mensagem_erro'))
        
        # Destaca linha com erro
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    
    # Salva em mem칩ria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retorna como download
    filename = f"inconsistencias_{empresa.get('cnpj', 'empresa')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== DASHBOARD - MONITOR RBT12 ====================
@app.get("/api/dashboard/metrics/{empresa_id}")
async def obter_metricas_rbt12(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    from dateutil.relativedelta import relativedelta
    
    # Verifica se a empresa pertence ao usu치rio
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inv치lido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa n칚o encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Calcula data de 12 meses atr치s
    hoje = datetime.utcnow()
    doze_meses_atras = hoje - relativedelta(months=12)
    
    # Pipeline de agrega칞칚o para somar faturamento dos 칰ltimos 12 meses
    pipeline = [
        {
            "$match": {
                "empresa_id": empresa_id,
                "data_emissao": {
                    "$gte": doze_meses_atras.isoformat(),
                    "$lte": hoje.isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": None,
                "faturamento_12_meses": {"$sum": "$valor_total"}
            }
        }
    ]
    
    resultado = await db.notas_fiscais.aggregate(pipeline).to_list(1)
    faturamento_atual = resultado[0]["faturamento_12_meses"] if resultado else 0.0
    
    # Obtem limite da empresa (padr칚o MEI: R$ 81.000,00)
    limite_anual = 81000.00  # Padr칚o MEI
    
    # Se houver limite cadastrado, usa ele
    if "limite_faturamento_anual" in empresa:
        limite_anual = float(empresa["limite_faturamento_anual"])
    elif empresa.get("regime_tributario") == "Simples Nacional":
        limite_anual = 4800000.00  # Limite Simples Nacional
    elif empresa.get("regime_tributario") == "Lucro Presumido":
        limite_anual = 78000000.00  # Limite Lucro Presumido
    
    # Calcula percentual de uso
    percentual_uso = (faturamento_atual / limite_anual * 100) if limite_anual > 0 else 0
    
    # Define status baseado no percentual
    if percentual_uso >= 100:
        status = "ESTOUROU"
    elif percentual_uso >= 80:
        status = "ALERTA"
    else:
        status = "OK"
    
    # Calcula quanto falta para o limite
    margem_disponivel = limite_anual - faturamento_atual
    
    return {
        "faturamento_atual": round(faturamento_atual, 2),
        "limite": round(limite_anual, 2),
        "percentual_uso": round(percentual_uso, 2),
        "status": status,
        "margem_disponivel": round(margem_disponivel, 2),
        "regime_tributario": empresa.get("regime_tributario", "MEI"),
        "razao_social": empresa.get("razao_social", "")
    }

# ==================== ROTA HOME ====================
@app.get("/")
async def home():
    return {
        "mensagem": "Fiscal F치cil API v2.0 - Sistema de Auditoria Fiscal 游",
        "status": "online"
    }

@app.get("/api/health")
async def health_check():
    try:
        await db.command("ping")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database error: {str(e)}")
